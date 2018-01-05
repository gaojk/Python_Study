# Author:Jing Lv
# 写excel

# 写2003 excel
import xlwt
# 写2007 excel
import openpyxl


def write03Excel(path, sheetName, value):
    wb = xlwt.Workbook(path)
    sheet = wb.add_sheet(sheetName)
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功")


def write07Excel(path, sheetName, value):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheetName

    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    wb.save(filename=path)
    print("写入数据成功")

value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]

file_2003 = '../dataFile/2003.xls'
file_2007 = '../dataFile/2007.xlsx'

write03Excel(file_2003, "book", value)
write07Excel(file_2007, "book", value)
